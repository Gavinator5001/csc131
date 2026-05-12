from __future__ import annotations

import csv
import json
import re
import zlib
from pathlib import Path

from civic_vote_scraper.minutes_db import MinutesDatabase


SHEET_TARGETS = {
    "a1": {
        "sheet_patterns": [r"schedule\s*a1\b", r"schedule\s*a-?1\b"],
        "type": "investment",
        "name_aliases": [
            "name of business entity",
        ],
    },
    "a2": {
        "sheet_patterns": [r"schedule\s*a-?2\b"],
        "type": "investment_or_entity",
        "name_aliases": [
            "name of business entity or trust",
            "name of business entity",
            "investment",
            "business entity/name",
        ],
    },
    "b": {
        "sheet_patterns": [r"schedule\s*b\b"],
        "type": "lender_or_property",
        "name_aliases": [
            "name of lender",
        ],
    },
    "c": {
        "sheet_patterns": [r"schedule\s*c\b"],
        "type": "income_source",
        "name_aliases": [
            "name of source",
        ],
    },
}

PDF_SCHEDULE_TARGETS = {
    "A1": {
        "section_patterns": [r"schedule\s*a-?1\b"],
        "record_type": "investment",
        "capture_labels": [
            r"name of business entity",
        ],
        "terminators": [
            r"general description of this business",
            r"fair market value",
            r"nature of investment",
            r"your business position",
            r"check one",
        ],
    },
    "A-2": {
        "section_patterns": [r"schedule\s*a-?2\b"],
        "record_type": "investment_or_entity",
        "capture_labels": [
            r"name of business entity or trust",
            r"name of business entity",
        ],
        "terminators": [
            r"general description of this business",
            r"fair market value",
            r"nature of investment",
            r"your business position",
            r"trust business entity",
            r"check one",
        ],
    },
    "B": {
        "section_patterns": [r"schedule\s*b\b"],
        "record_type": "lender_or_property",
        "capture_labels": [
            r"name of lender",
        ],
        "terminators": [
            r"address of real property",
            r"assessor'?s parcel number",
            r"fair market value",
            r"if additional space is needed",
            r"check one",
        ],
    },
    "C": {
        "section_patterns": [r"schedule\s*c\b"],
        "record_type": "income_source",
        "capture_labels": [
            r"name of source of income",
        ],
        "terminators": [
            r"address \(business address acceptable\)",
            r"business activity, if any",
            r"your business position",
            r"gross income received",
            r"consideration for which income was received",
            r"check one",
        ],
    },
}

COMMON_SECTION_TERMINATORS = [
    r"official use only",
    r"statement of economic interests",
    r"page\s*[\-]?\s*\d+",
    r"fppc form 700",
    r"schedule\s*a-?1\b",
    r"schedule\s*a-?2\b",
    r"schedule\s*b\b",
    r"schedule\s*c\b",
]

NON_DATA_VALUES = {
    "", "n/a", "na", "none", "not applicable",
    "schedule attached", "continued", "continued on next page"
}

PDF_SECTION_STOP_LINES = [
    r"expanded statement list",
    r"filing info",
    r"filing data",
    r"gifts and travel",
    r"schedule\s*a-?1\b",
    r"schedule\s*a-?2\b",
    r"schedule\s*b\b",
    r"schedule\s*c\b",
]

PDF_FILER_LABELS = [
    r"first name",
    r"middle name",
    r"middle initial",
    r"last name",
    r"surname",
    r"agency",
    r"position",
]

PDF_FILING_INFO_LABELS = [
    r"filing type",
    r"filing year",
    r"due date",
    r"filed date",
]

RAW_PDF_JUNK_VALUES = {
    "adobe",
    "ucs",
    "c",
    "m",
    "b",
    "san fppc",
    "san: fppc",
}

RAW_PDF_GENERIC_VALUES = {
    "holding company",
    "media",
    "internet",
    "computer",
    "auto rental",
    "consumer product sales",
    "banking services",
    "bank",
    "real estate development",
    "financial",
    "healthcare",
    "communications",
    "communication",
    "media communications",
    "energy company",
    "software internet",
    "baseball shares",
    "financial stock",
    "marketing automation software",
    "flowers",
    "etf",
    "etf sector",
    "hotels",
    "co owner",
}

PDF_DESCRIPTION_SIGNAL_PATTERN = re.compile(
    r"\b("
    r"technology|financial services?|communication services?|medical devices|healthcare|"
    r"pharmaceuticals?|beverages?|retail sales|real estate|utilities|energy|food|"
    r"packaged foods?|specialty chemicals?|specialty retail|household products?|"
    r"railroads?|restaurants?|solar technology|industrials?|tools and accessories|"
    r"internet retail|electric and gas utility|retirement pension|medical|devices|"
    r"communications?|services|chemicals?|beverage|retail|solar|pension"
    r")\b",
    re.I,
)

PDF_INSTITUTION_SIGNAL_PATTERN = re.compile(
    r"\b("
    r"inc|corp|corporation|company|group|holdings?|bank|bancorp|trust|fund|partners?|"
    r"llc|ltd|class|etf|global|capital|ventures|international|"
    r"laborator(?:y|ies)|administration|association|"
    r"systems?|automotive|scientific|wholesale|railroad|board|commission|authority|"
    r"agency|department|library|district|county|city|college|university|school|"
    r"government|hospital|insurance|transit|logistics|biosciences?|solar"
    r")\b",
    re.I,
)

PDF_ROLE_STATUS_PATTERN = re.compile(
    r"\b("
    r"retiree|commissioner|supervisor|trustee|board member|member|alternate members?|"
    r"director|manager|judge|attorney|counsel|candidate|chief|officer|administrator"
    r")\b",
    re.I,
)

ENTITY_LABEL_PREFIX_PATTERNS = [
    r"name of source of income(?: \(not an acronym\))?",
    r"name of source(?: \(not an acronym\))?",
    r"name of business entity or trust",
    r"name of business entity",
    r"name of lender",
    r"address \(business or agency address recommended - public document\)",
    r"address \(business address acceptable(?: recommended - public document)?\)",
    r"general description of this business",
    r"business activity,? if any,? of (?:source|lender)",
    r"your business position",
    r"office, agency, or court",
    r"agency name \(do not use acronyms\)",
    r"print name office, agency or court",
]

FORM700_BOILERPLATE_PATTERNS = [
    r"\bbusiness or agency address recommended\b",
    r"\bbusiness address acceptable\b",
    r"\bcheck one(?: circle)?\b",
    r"\bcheck at least one box\b",
    r"\bdo not use acronyms\b",
    r"\bfor self-employed use schedule a-?2\b",
    r"\bif gift\b",
    r"\bif applicable\b",
    r"\bdescribe\b",
    r"\bpublic document\b",
    r"\bofficial use only\b",
    r"\bstatement of economic interests\b",
    r"\bfppc form 700\b",
    r"\bgross income received\b",
    r"\bconsideration for which income was received\b",
    r"\bhighest balance during reporting period\b",
    r"\binterest rate term\b",
    r"\bloan repayment\b",
    r"\bguarantor\b",
    r"\baddress of real property\b",
    r"\bassessor'?s parcel number\b",
    r"\bprint name office, agency or court\b",
    r"\bagency name\b.*\bdo not use acronyms\b",
    r"\bannual: the period covered is\b",
    r"\bassuming office: date assumed\b",
    r"\bleaving office\b",
    r"\bcandidate\b",
    r"\bschedule attached\b",
    r"\bgeneral description of this business\b",
    r"\byour business position\b",
    r"\bbusiness activity,? if any,? of (?:source|lender)\b",
    r"\bname of (?:source(?: of income)?|business entity(?: or trust)?|lender)\b",
    r"\bfiled date\b",
    r"\bpage\s+\d+\b",
    r"\bownership interest is\b",
    r"\bnone personal residence\b",
    r"\bcover page attachment\b",
    r"\bexpanded statement list\b",
    r"\bsecurity for loan\b",
    r"\bloans received or outstanding during the reporting period\b",
    r"\bfair political practices commission\b",
    r"\bstate judge\b",
    r"\bcourt commissioner\b",
    r"\bvalue description of gift\b",
    r"\bfair market value\b",
    r"\byrs?\. remaining\b",
    r"\bsan:\s*[0-9a-z\-]+\b",
    r"\bproperty ownership/deed of trust\b",
]

FORM700_BOILERPLATE_EXACT_KEYS = {
    "agency",
    "assets",
    "authority",
    "comments",
    "guarantor",
    "other",
    "positions",
    "schedule c",
    "security for loan",
    "salary spouses or registered domestic partners income",
    "fair political practices commission",
    "california form",
    "name of lender",
    "name of source not an acronym",
    "name of business entity",
    "name",
    "first",
    "middle",
    "last",
    "surname",
    "describe",
    "state",
    "tickets",
    "sale of",
    "sonoma",
    "county of sonoma",
    "commissioner",
    "cover page attachment",
    "expanded statement list",
    "none personal residence",
    "office agency or court",
    "general description of this business",
    "general description of",
    "alifornia form",
    "alifornia form 700",
    "omments",
    "over page attachment",
    "oans received or outstanding during the reporting period",
    "nterest rate term monthsyears",
    "f applicable list date",
    "ommissioner",
    "ounty of sonoma",
    "nterest rate term months/years",
    "page 13",
    "page 15",
    "jurisdiction",
    "statement",
    "description of business activity or",
    "trust go to 2 business entity complete the box then go to 2",
    "type of",
    "easehold",
    "ommission lafco",
    "onservation and",
    "nvestments",
    "amendment",
    "eased by the business entity or trust",
    "none noneor or names listed below names listed below",
    "ownership/deed of trust easement",
    "of business entities/trusts",
    "ncome r eceived",
    "business entity or trust 1 business entity or trust",
    "acqu ired d isposed",
    "natu re of interest",
    "nature of interest",
}

FORM700_BOILERPLATE_COMPACT_FRAGMENTS = {
    "coverpageattachment",
    "expandedstatementlist",
    "nonepersonalresidence",
    "saleof",
    "san081800216sth0216",
    "californiaform700",
    "loansreceivedoroutstandingduringthereportingperiod",
    "businessactivityifanyoflender",
    "businessactivityifanyofsource",
    "datemmddyyvaluedescriptionofgift",
    "fairpoliticalpracticescommission",
    "yourbusinessposition",
    "securityforloan",
    "fairmarketvalue",
    "highestbalanceduringreportingperiod",
    "propertyownershipdeedoftrust",
    "yrsremaining",
    "nameofsourcenotanacronym",
    "nameofbusinessentity",
    "officeagencyorcourt",
    "generaldescriptionofthisbusiness",
    "aliforniaform",
    "comments",
    "overpageattachment",
    "oansreceivedoroutstandingduringthereportingperiod",
    "interestrateterm",
    "fapplicablelistdate",
    "commissioner",
    "over100000",
    "nterestrateterm",
    "descriptionofbusinessactivityor",
    "trustgoto2businessentitycompletetheboxthengoto2",
    "page13",
    "page15",
    "natureofinvestment",
    "incomereceived",
    "easehold",
    "ownershipdeedoftrusteasement",
    "ofbusinessentitiestrusts",
    "nonenoneorornameslistedbelownameslistedbelow",
    "easedbythebusinessentityortrust",
    "ncomereceived",
    "businessentityortrust1businessentityortrust",
    "acquireddisposed",
    "natureofinterest",
}


def norm_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def norm_key(value):
    text = norm_text(value).lower().replace("&", "and")
    text = re.sub(r"[^a-z0-9\s\-/]", "", text)
    return re.sub(r"\s+", " ", text).strip()


def _looks_like_name_initial(value: str) -> bool:
    text = norm_text(value).rstrip(".")
    return bool(re.fullmatch(r"[A-Za-z]", text))


def _is_raw_pdf_junk_line(value: str) -> bool:
    text = norm_text(value)
    key = norm_key(text)
    if key not in RAW_PDF_JUNK_VALUES:
        return False
    if key in {"b", "c", "m"} and _looks_like_name_initial(text):
        return False
    return True


def split_person_name(value: str) -> tuple[str, str, str]:
    text = norm_text(value)
    if not text:
        return "", "", ""
    if "," in text:
        last, rest = [part.strip() for part in text.split(",", 1)]
        rest_parts = rest.split()
        first = rest_parts[0] if rest_parts else ""
        middle = " ".join(rest_parts[1:]) if len(rest_parts) > 1 else ""
        return first, middle, last
    parts = text.split()
    if len(parts) == 1:
        return parts[0], "", ""
    if len(parts) == 2:
        return parts[0], "", parts[1]
    if len(parts) == 3:
        if _looks_like_name_initial(parts[1]):
            return parts[0], parts[1].rstrip("."), parts[2]
        if _looks_like_name_initial(parts[2]):
            return parts[1], parts[2].rstrip("."), parts[0]
    return parts[0], " ".join(parts[1:-1]), parts[-1]


def find_sheet(wb, target_key):
    cfg = SHEET_TARGETS[target_key]
    for ws in wb.worksheets:
        title = norm_key(ws.title)
        for pattern in cfg["sheet_patterns"]:
            if re.search(pattern, title):
                return ws
    return None


def score_header_row(row_values, aliases):
    keys = [norm_key(v) for v in row_values if norm_text(v)]
    score = 0
    for alias in aliases:
        alias_key = norm_key(alias)
        if alias_key in keys:
            score += 3
    return score


def detect_header(ws, aliases, max_scan_rows=20):
    best_row, best_score, best_headers = None, -1, None
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        score = score_header_row(values, aliases)
        if score > best_score:
            best_row, best_score, best_headers = r, score, values
    if best_row is None or best_score <= 0:
        raise RuntimeError(f"Could not detect header row in sheet '{ws.title}'")
    return best_row, [norm_text(v) for v in best_headers]


def build_header_index(headers):
    idx = {}
    for i, h in enumerate(headers, start=1):
        key = norm_key(h)
        if key and key not in idx:
            idx[key] = i
    return idx


def pick_name_column(header_index, aliases):
    alias_keys = [norm_key(a) for a in aliases]
    for alias in alias_keys:
        if alias in header_index:
            return header_index[alias], alias
    return None, None


def row_to_dict(ws, row_num, headers):
    out = {}
    for i, header in enumerate(headers, start=1):
        out[header or f"column_{i}"] = norm_text(ws.cell(row_num, i).value)
    return out


def is_probable_data_row(values):
    nonempty = [v for v in values if norm_text(v)]
    return len(nonempty) >= 2


def extract_form700_owner(wb):
    ws = wb.worksheets[0]
    owner_last_name = norm_text(ws.cell(1, 1).value)
    owner_first_name = norm_text(ws.cell(1, 2).value)
    owner_full_name = f"{owner_first_name} {owner_last_name}".strip()
    print(f"[info] owner from workbook cols 1-2: {owner_full_name or '(blank)'}")
    return {
        "owner_last_name": owner_last_name,
        "owner_first_name": owner_first_name,
        "owner_middle_name": "",
        "owner_full_name": owner_full_name,
        "filer_position_title": "",
        "filer_agency_name": "",
        "filer_entity_name": "",
    }


def extract_sheet_records(ws, target_key, owner_meta):
    cfg = SHEET_TARGETS[target_key]
    header_row, headers = detect_header(ws, cfg["name_aliases"])
    header_index = build_header_index(headers)
    name_col, matched_header = pick_name_column(header_index, cfg["name_aliases"])
    if not name_col:
        raise RuntimeError(f"Could not identify strict name column in sheet '{ws.title}'")
    print(f"[info] using entity column '{matched_header}' in sheet '{ws.title}'")

    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        text_values = [norm_text(v) for v in values]
        if not any(text_values):
            continue
        if not is_probable_data_row(text_values):
            continue
        raw_name = norm_text(ws.cell(r, name_col).value)
        if norm_key(raw_name) in NON_DATA_VALUES or not raw_name:
            continue

        rec = row_to_dict(ws, r, headers)
        rec["_sheet"] = ws.title
        rec["_schedule"] = target_key.upper().replace("A2", "A-2")
        rec["_record_type"] = cfg["type"]
        rec["_name_column"] = matched_header
        rec["_row_number"] = r
        rec["entity_name"] = raw_name
        rec["raw_value"] = raw_name
        rec["owner_last_name"] = owner_meta["owner_last_name"]
        rec["owner_first_name"] = owner_meta["owner_first_name"]
        rec["owner_middle_name"] = owner_meta.get("owner_middle_name", "")
        rec["owner_full_name"] = owner_meta["owner_full_name"]
        rec["filer_position_title"] = owner_meta.get("filer_position_title", "")
        rec["filer_agency_name"] = owner_meta.get("filer_agency_name", "")
        rec["filer_entity_name"] = owner_meta.get("filer_entity_name", "")
        rec["jurisdiction"] = owner_meta.get("jurisdiction", "")
        rec["_source_pdf_path"] = owner_meta.get("_source_pdf_path", "")
        records.append(rec)
    return records


def parse_form700_workbook(input_path: str | Path):
    from openpyxl import load_workbook

    input_path = Path(input_path)
    print(f"[info] opening Form 700 workbook: {input_path}")
    wb = load_workbook(input_path, data_only=True, read_only=True)
    print(f"[info] workbook loaded with {len(wb.worksheets)} sheets")

    owner_meta = extract_form700_owner(wb)

    all_records = []
    for target_key in ["a1", "a2", "b", "c"]:
        schedule_label = target_key.upper().replace("A2", "A-2")
        print(f"[info] looking for schedule {schedule_label}")
        ws = find_sheet(wb, target_key)

        if ws is None:
            print(f"[info] schedule {schedule_label} not found")
            continue

        print(f"[info] found sheet: {ws.title}")
        rows = extract_sheet_records(ws, target_key, owner_meta)
        print(f"[info] extracted {len(rows)} rows from {ws.title}")
        all_records.extend(rows)

    print(f"[info] Form 700 workbook parsing complete: {len(all_records)} total rows")
    return all_records


def _decode_pdf_literal_bytes(data: bytes) -> str:
    out = []
    index = 0
    while index < len(data):
        byte = data[index]
        if byte == 92 and index + 1 < len(data):
            index += 1
            escaped = data[index]
            mapping = {
                110: "\n",
                114: "\r",
                116: "\t",
                98: "\b",
                102: "\f",
                40: "(",
                41: ")",
                92: "\\",
            }
            if escaped in mapping:
                out.append(mapping[escaped])
            elif 48 <= escaped <= 55:
                oct_digits = bytes([escaped])
                for _ in range(2):
                    if index + 1 < len(data) and 48 <= data[index + 1] <= 55:
                        index += 1
                        oct_digits += bytes([data[index]])
                    else:
                        break
                out.append(chr(int(oct_digits, 8)))
            else:
                out.append(chr(escaped))
        else:
            out.append(chr(byte))
        index += 1
    return "".join(out)


def _looks_like_human_pdf_text(value: str) -> bool:
    text = value.strip()
    if not text or len(text) > 200:
        return False
    printable = sum(32 <= ord(ch) <= 126 or ch in "\n\r\t" for ch in text)
    return printable / max(len(text), 1) >= 0.85


def _iter_pdf_stream_payloads(pdf_bytes: bytes):
    for match in re.finditer(rb"<<(.*?)>>\s*stream\r?\n", pdf_bytes, re.S):
        stream_dict = match.group(1)
        start = match.end()
        end = pdf_bytes.find(b"endstream", start)
        if end == -1:
            continue
        payload = pdf_bytes[start:end].rstrip(b"\r\n")
        if b"FlateDecode" in stream_dict:
            try:
                payload = zlib.decompress(payload)
            except Exception:
                continue
        yield payload


def _extract_pdf_literal_strings(payload: bytes) -> list[str]:
    strings = []
    index = 0
    while index < len(payload):
        if payload[index] != 40:
            index += 1
            continue

        index += 1
        depth = 1
        literal = bytearray()
        while index < len(payload) and depth > 0:
            byte = payload[index]
            if byte == 92 and index + 1 < len(payload):
                literal.append(byte)
                index += 1
                literal.append(payload[index])
            elif byte == 40:
                depth += 1
                literal.append(byte)
            elif byte == 41:
                depth -= 1
                if depth > 0:
                    literal.append(byte)
            else:
                literal.append(byte)
            index += 1

        text = norm_text(_decode_pdf_literal_bytes(bytes(literal)))
        if _looks_like_human_pdf_text(text):
            strings.append(text)
    return strings


def _extract_pdf_title_from_bytes(pdf_bytes: bytes) -> str:
    match = re.search(rb"/Title\s*\((.*?)\)", pdf_bytes, re.S)
    if not match:
        return ""
    return norm_text(_decode_pdf_literal_bytes(match.group(1)))


def _extract_pdf_text_and_title_from_raw_bytes(input_path: Path) -> tuple[str, str]:
    pdf_bytes = input_path.read_bytes()
    title = _extract_pdf_title_from_bytes(pdf_bytes)
    parts = []
    for payload in _iter_pdf_stream_payloads(pdf_bytes):
        parts.extend(_extract_pdf_literal_strings(payload))
    return "\n".join(parts), title


def extract_pdf_text(input_path: str | Path) -> str:
    text, _title = extract_pdf_text_and_title(input_path)
    return text


def extract_pdf_text_and_title(input_path: str | Path) -> tuple[str, str]:
    path = Path(input_path)
    errors = []

    try:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        parts = []
        for page in reader.pages:
            parts.append(page.extract_text() or "")

        metadata = reader.metadata or {}
        title = norm_text(getattr(metadata, "title", "") or metadata.get("/Title", ""))
        text = "\n".join(parts)
        if norm_text(text) or title:
            return text, title
    except Exception as exc:
        errors.append(f"pypdf: {exc}")

    try:
        import pdfplumber

        parts = []
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                parts.append(page.extract_text() or "")
            metadata = pdf.metadata or {}
        title = norm_text(metadata.get("Title", "") or metadata.get("/Title", ""))
        text = "\n".join(parts)
        if norm_text(text) or title:
            return text, title
    except Exception as exc:
        errors.append(f"pdfplumber: {exc}")

    text, title = _extract_pdf_text_and_title_from_raw_bytes(path)
    if norm_text(text) or title:
        if errors:
            print(f"[info] Form 700 PDF extraction fallback used for {path.name}: {'; '.join(errors)}")
        return text, title

    detail = "; ".join(errors) if errors else "no extractor returned usable content"
    raise RuntimeError(f"Could not extract text from PDF {path}: {detail}")


def _clean_pdf_candidate(value: str) -> str:
    text = norm_text(value)
    text = re.sub(r"\s*\(\s*continued\s*\)\s*$", "", text, flags=re.I)
    text = re.sub(r"[_\[\]\u25a0\u25a1]+", " ", text)
    text = re.sub(r"^[\"'`“”‘’•*►\-\s]+", "", text)
    text = re.sub(r"\s{2,}", " ", text).strip(" -:;,")
    while True:
        original = text
        text = re.sub(r"^(?:(?:\d+|[ivxlc]{1,6})(?:[.)*\-]|\s)\s*)+", "", text, flags=re.I)
        for pattern in ENTITY_LABEL_PREFIX_PATTERNS:
            text = re.sub(rf"^(?:{pattern})\s*[:\-–—]?\s*", "", text, flags=re.I)
        text = re.sub(r"^[\"'`“”‘’•*►\-\s?]+", "", text).strip(" -:;,")
        if text == original:
            break
    return text


def _looks_like_form700_boilerplate(value: str) -> bool:
    text = norm_text(value)
    key = norm_key(text)
    compact_key = re.sub(r"[^a-z0-9]+", "", key)
    if not text or not key:
        return True
    if key in FORM700_BOILERPLATE_EXACT_KEYS:
        return True
    if compact_key in FORM700_BOILERPLATE_COMPACT_FRAGMENTS:
        return True
    if any(fragment in compact_key for fragment in FORM700_BOILERPLATE_COMPACT_FRAGMENTS):
        return True
    if len(text) > 180:
        return True
    if len(text.split()) > 18:
        return True
    if _is_contact_or_address_line(text):
        return True
    if re.fullmatch(r"[\W\d_ ]+", text):
        return True
    if re.search(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", text) and len(text.split()) > 4:
        return True
    if re.search(r"\b(?:annual|assuming office|leaving office|candidate)\b", key) and len(text.split()) > 3:
        return True
    if re.search(r"\b(?:0|o|el|li)\s+(?:none|over|\$)", key):
        return True
    if re.search(r"\b(?:schedule|fppc|public document|verification)\b", key) and len(text.split()) <= 8:
        return True
    if (
        "agency" in key
        and re.search(r"addre[a-z]*", key)
        and re.search(r"recommend[a-z]*", key)
        and "public" in key
        and re.search(r"docu[a-z]*", key)
    ):
        return True
    if (
        ("business" in key or "source" in key or "lender" in key)
        and re.search(r"addre[a-z]*", key)
        and re.search(r"accept[a-z]*|recommend[a-z]*", key)
    ):
        return True
    if re.fullmatch(r"(?:first|middle|last|surname)(?:\s+(?:first|middle|last|surname))*", key):
        return True
    if re.fullmatch(r"(?:describe|if gift|ownership interest.*)", key):
        return True
    if re.fullmatch(r"(?:name of .+|general description.+|business activity.+|address.+)", key):
        return True
    if re.fullmatch(r"(?:board member county of|authority board member county of)", key):
        return True
    if re.fullmatch(r"(?:name|commissioner|court|judge)", key):
        return True
    if re.fullmatch(r".*\bcounty of\b", key):
        return True
    if re.fullmatch(r".*\bover\s*\$?\d[\d,]*\b.*", key):
        return True
    for pattern in FORM700_BOILERPLATE_PATTERNS:
        if re.search(pattern, text, flags=re.I):
            return True
    return False


def _is_valid_pdf_entity(value: str) -> bool:
    text = _clean_pdf_candidate(value)
    key = norm_key(text)
    if not key or key in NON_DATA_VALUES:
        return False
    if len(key) < 3:
        return False
    if re.fullmatch(r"(yes|no|unknown|none|check one)", key):
        return False
    if _looks_like_form700_boilerplate(text):
        return False
    return True


def _sanitize_context_field(value: str) -> str:
    text = norm_text(value)
    if not text:
        return ""
    key = norm_key(text)
    if not key:
        return ""
    if _looks_like_form700_boilerplate(text):
        return ""
    if re.search(
        r"\b(type of statement|check at least one box|agency name|do not use acronyms|california form 700|fair political practices commission|state judge|court commissioner|official use only)\b",
        text,
        flags=re.I,
    ):
        return ""
    return text


def _owner_meta_from_filing_metadata(filing_metadata: dict | None, source_pdf_path: str) -> dict:
    filing_metadata = filing_metadata or {}
    first = norm_text(filing_metadata.get("filer_first_name", ""))
    middle = norm_text(filing_metadata.get("filer_middle_name", ""))
    last = norm_text(filing_metadata.get("filer_last_name", ""))
    full = norm_text(filing_metadata.get("filer_full_name", "")) or " ".join(
        [part for part in [first, middle, last] if part]
    ).strip()
    return {
        "owner_last_name": last,
        "owner_first_name": first,
        "owner_middle_name": middle,
        "owner_full_name": full,
        "filer_position_title": norm_text(filing_metadata.get("position_title", "")),
        "filer_agency_name": norm_text(filing_metadata.get("agency_name", "")),
        "filer_entity_name": norm_text(filing_metadata.get("entity_name", "")),
        "jurisdiction": norm_text(filing_metadata.get("jurisdiction", "")),
        "_source_pdf_path": source_pdf_path,
    }


def _extract_name_from_pdf_title(title: str) -> tuple[str, str, str]:
    clean = norm_text(title)
    if not clean:
        return "", "", ""

    match = re.match(r"^(.*?)\s+statement of economic interests\b", clean, flags=re.I)
    if match:
        clean = norm_text(match.group(1))

    clean = re.sub(r"\s*-\s*redacted\s*$", "", clean, flags=re.I)
    clean = re.sub(r"\s+form\s*700.*$", "", clean, flags=re.I)
    if not clean:
        return "", "", ""

    return split_person_name(clean)


def _normalized_lines(text: str) -> list[str]:
    return [norm_text(line) for line in str(text or "").splitlines() if norm_text(line)]


def _extract_named_section_lines(
    lines: list[str],
    title_patterns: list[str],
    required_patterns: list[str],
    stop_patterns: list[str],
) -> list[str]:
    for index, line in enumerate(lines):
        if not any(re.fullmatch(pattern, line, flags=re.I) for pattern in title_patterns):
            continue
        window = "\n".join(lines[index : index + 16])
        if required_patterns and not any(re.search(pattern, window, flags=re.I) for pattern in required_patterns):
            continue

        section = [lines[index]]
        for next_line in lines[index + 1 :]:
            if any(re.fullmatch(pattern, next_line, flags=re.I) for pattern in stop_patterns):
                break
            section.append(next_line)
        return section
    return []


def _extract_line_labeled_value(lines: list[str], label_patterns: list[str], all_label_patterns: list[str]) -> str:
    for index, line in enumerate(lines):
        for pattern in label_patterns:
            match = re.match(rf"^(?:{pattern})\s*:?\s*(.*)$", line, flags=re.I)
            if not match:
                continue

            same_line = _clean_pdf_candidate(match.group(1))
            if same_line and not any(re.fullmatch(other, same_line, flags=re.I) for other in all_label_patterns):
                return same_line

            for next_line in lines[index + 1 :]:
                if any(re.fullmatch(other, next_line, flags=re.I) for other in all_label_patterns):
                    break
                candidate = _clean_pdf_candidate(next_line)
                if candidate:
                    return candidate
            break
    return ""


def _extract_text_labeled_value(text: str, label_patterns: list[str], stop_patterns: list[str]) -> str:
    collapsed = "\n".join(_normalized_lines(text))
    if not collapsed:
        return ""

    stop_pattern = "|".join(f"(?:{pattern})" for pattern in stop_patterns)
    for pattern in label_patterns:
        match = re.search(
            rf"(?:{pattern})\s*:?\s*(.+?)(?=(?:{stop_pattern})\s*:?\s*|$)",
            collapsed,
            flags=re.I | re.S,
        )
        if not match:
            continue
        candidate = _clean_pdf_candidate(match.group(1))
        if candidate:
            return candidate
    return ""


def _is_contact_or_address_line(value: str) -> bool:
    text = norm_text(value)
    key = norm_key(text)
    if not key:
        return True
    if "@" in text:
        return True
    if re.fullmatch(r"[\d\s().\-+/]+", text):
        return True
    if re.search(r"\b\d{5}(?:-\d{4})?\b", text):
        return True
    if re.search(r"\b(?:street|st|avenue|ave|road|rd|drive|dr|boulevard|blvd|lane|ln|suite|ste|room|dept|department)\b", key):
        return True
    if re.search(r"\bp\.?\s*o\.?\s*box\b", key):
        return True
    if key in {"ca", "california"}:
        return True
    return False


def _extract_unstructured_header_fields(lines: list[str], fallback: dict) -> tuple[str, str]:
    if not lines:
        return "", ""

    boundary = len(lines)
    for index, line in enumerate(lines):
        if re.search(r"\bfiled date\b", line, flags=re.I) or re.search(r"\bsan:\s*fppc\b", line, flags=re.I):
            boundary = index
            break

    header_lines = [line for line in lines[:boundary] if not _is_raw_pdf_junk_line(line)]
    if not header_lines:
        return "", ""

    owner_bits = {
        norm_key(fallback.get("owner_full_name", "")),
        norm_key(fallback.get("owner_first_name", "")),
        norm_key(fallback.get("owner_last_name", "")),
    }

    position = ""
    agency = ""
    agency_index = None

    preferred_patterns = [
        r"\b(county|city|district|agency|commission|authority|court|department)\b",
        r"\b(school|library|board|supervisor)\b",
    ]
    for pattern in preferred_patterns:
        for index, line in enumerate(header_lines):
            key = norm_key(line)
            if not key or key in owner_bits or _is_contact_or_address_line(line):
                continue
            if len(line) <= 2:
                continue
            if re.search(pattern, line, flags=re.I):
                agency = line
                agency_index = index
                break
        if agency_index is not None:
            break

    if agency_index is not None and agency_index > 0:
        candidate = header_lines[agency_index - 1]
        if norm_key(candidate) not in owner_bits and not _is_contact_or_address_line(candidate) and len(candidate) > 2:
            position = candidate

    return position, agency


def _looks_like_cover_name_line(value: str) -> bool:
    text = norm_text(value)
    key = norm_key(text)
    if not text or not key or key in RAW_PDF_JUNK_VALUES or key in RAW_PDF_GENERIC_VALUES:
        return False
    if _is_contact_or_address_line(text):
        return False
    if re.search(
        r"\b(county|city|district|agency|commission|authority|court|department|board|supervisor|judge|form|statement|filed|date|attached|schedule|sonoma)\b",
        key,
    ):
        return False
    parts = [re.sub(r"[^A-Za-z'\-]", "", part) for part in text.split()]
    parts = [part for part in parts if part]
    if not parts or len(parts) > 3:
        return False
    if all(len(part) == 1 for part in parts):
        return False
    return all(part[:1].isupper() and part[1:].islower() for part in parts if len(part) > 1)


def _looks_like_cover_middle_line(value: str) -> bool:
    text = norm_text(value)
    if not text:
        return False
    if _looks_like_name_initial(text):
        return True
    return _looks_like_cover_name_line(text)


def _extract_unstructured_cover_owner_fields(
    lines: list[str],
    fallback: dict,
) -> tuple[str, str, str, str, str]:
    if not lines:
        return "", "", "", "", ""

    boundary = len(lines)
    for index, line in enumerate(lines):
        if re.search(r"\bfiled date\b", line, flags=re.I) or re.search(r"\bsan:\s*", line, flags=re.I):
            boundary = index
            break

    header_lines = [line for line in lines[:boundary] if not _is_raw_pdf_junk_line(line)]
    if not header_lines:
        return "", "", "", "", ""

    tail = header_lines[max(0, len(header_lines) - 24) :]
    agency_signal = re.compile(r"\b(county|city|district|agency|commission|authority|court|department|board|library)\b", re.I)
    board_signal = re.compile(r"\b(board|department|district|division|commission|authority|library|court)\b", re.I)
    stop_signal = re.compile(r"\b(see attached|schedule attached|filed date|san:)\b", re.I)

    for index in range(len(tail) - 3, -1, -1):
        last = norm_text(tail[index])
        first = norm_text(tail[index + 1]) if index + 1 < len(tail) else ""
        third_line = norm_text(tail[index + 2]) if index + 2 < len(tail) else ""
        fourth_line = norm_text(tail[index + 3]) if index + 3 < len(tail) else ""
        if not (_looks_like_cover_name_line(last) and _looks_like_cover_name_line(first)):
            continue

        middle = ""
        agency_line = third_line
        cursor = index + 3
        if _looks_like_cover_middle_line(third_line) and agency_signal.search(fourth_line):
            middle = third_line.rstrip(".")
            agency_line = fourth_line
            cursor = index + 4

        if not agency_signal.search(agency_line):
            continue

        agency_parts = [agency_line]
        while cursor < len(tail):
            line = norm_text(tail[cursor])
            if not line or stop_signal.search(line):
                break
            if board_signal.search(line):
                agency_parts.append(line)
                cursor += 1
                continue
            break

        position = ""
        while cursor < len(tail):
            line = norm_text(tail[cursor])
            key = norm_key(line)
            cursor += 1
            if not line or _is_raw_pdf_junk_line(line) or key in RAW_PDF_GENERIC_VALUES:
                continue
            if stop_signal.search(line) or _is_contact_or_address_line(line):
                break
            if re.fullmatch(r"\d+", line):
                continue
            position = line
            break

        agency = " - ".join(part for part in agency_parts if part)
        return first, middle, last, position, agency

    return "", "", "", "", ""


def _looks_like_description_line(value: str) -> bool:
    text = norm_text(value)
    key = norm_key(text)
    if not key:
        return False
    if key in RAW_PDF_GENERIC_VALUES:
        return True
    if re.fullmatch(
        r"(technology|financial services|communication services|medical devices|healthcare|"
        r"pharmaceuticals?|beverages?|retail sales|real estate|utilities|energy|food|"
        r"packaged foods?|specialty chemicals?|specialty retail|household products?|"
        r"railroads?|restaurants?|solar technology|industrials?|tools and accessories|"
        r"internet retail|electric and gas utility|retirement pension)",
        key,
    ):
        return True
    if PDF_DESCRIPTION_SIGNAL_PATTERN.search(text) and not PDF_INSTITUTION_SIGNAL_PATTERN.search(text):
        return True
    if re.search(
        r"\b(services|shares|rental|travel|property|investment|income|owner|liability|testing|automation|sales|development|broadcasting)\b",
        key,
    ) and not re.search(
        r"\b(inc|corp|corporation|company|group|holdings?|bank|bancorp|trust|fund|partners?|llc|ltd|class|etf|global|energy|capital|ventures|communications?|pharmaceuticals?)\b",
        key,
    ):
        return True
    if "company" in key and sum(token[:1].isupper() for token in text.split() if token[:1].isalpha()) <= 1:
        return True
    if text.lower() == text and len(text.split()) <= 8:
        return True
    if re.fullmatch(r"[a-z\s\-&,/]+", text) and len(text.split()) <= 6:
        return True
    if len(text.split()) >= 4 and sum(token[:1].isupper() for token in text.split() if token[:1].isalpha()) <= 1:
        return True
    return False


def _entity_signal_score(value: str) -> int:
    text = norm_text(value)
    key = norm_key(text)
    if not key or key in RAW_PDF_JUNK_VALUES:
        return -10
    if key in RAW_PDF_GENERIC_VALUES:
        return -5
    if _is_contact_or_address_line(text):
        return -10

    tokens = re.findall(r"[A-Za-z0-9&'.-]+", text)
    alpha_tokens = [token for token in tokens if re.search(r"[A-Za-z]", token)]
    if not alpha_tokens:
        return -10

    score = 0
    if len(alpha_tokens) >= 2:
        score += 1
    capitals = sum(token[:1].isupper() for token in alpha_tokens if token[:1].isalpha())
    if capitals >= max(1, len(alpha_tokens) // 2):
        score += 1
    if len(alpha_tokens) == 1 and alpha_tokens[0][:1].isupper() and len(alpha_tokens[0]) >= 4:
        score += 1
    if re.search(
        r"\b(inc|corp|corporation|company|group|holdings?|bank|bancorp|trust|fund|partners?|llc|ltd|class|etf|global|energy|capital|ventures|communications?|pharmaceuticals?|international)\b",
        key,
    ):
        score += 2
    if re.search(r"\bclass\s+[a-z0-9]+\b", key):
        score += 1
    if text.lower() == text:
        score -= 1
    return score


def _looks_like_role_or_status_line(value: str) -> bool:
    text = norm_text(value)
    if not text:
        return False
    return bool(PDF_ROLE_STATUS_PATTERN.search(text))


def _has_business_or_institution_signal(value: str) -> bool:
    text = norm_text(value)
    if not text:
        return False
    return bool(PDF_INSTITUTION_SIGNAL_PATTERN.search(text))


def _looks_like_person_name_line(value: str) -> bool:
    text = norm_text(value)
    if not text or text.startswith("The "):
        return False
    key = norm_key(text)
    if _looks_like_description_line(text):
        return False
    if _looks_like_role_or_status_line(text):
        return False
    if _has_business_or_institution_signal(text):
        return False
    if re.search(
        r"\b(inc|corp|corporation|company|group|holdings?|bank|bancorp|trust|fund|partners?|llc|ltd|class|etf|global|energy|capital|ventures|communications?|pharmaceuticals?)\b",
        key,
    ):
        return False
    if "@" in text:
        return False
    parts = [part for part in re.split(r"\s+", text) if part]
    alpha_parts = [part for part in parts if re.search(r"[A-Za-z]", part)]
    if len(alpha_parts) < 2 or len(alpha_parts) > 6:
        return False
    cleaned = [re.sub(r"[^A-Za-z'\-]", "", part) for part in alpha_parts if re.sub(r"[^A-Za-z'\-]", "", part)]
    if not cleaned:
        return False
    if "and" in {part.lower() for part in cleaned}:
        return True
    return all(part[:1].isupper() and part[1:].islower() for part in cleaned if len(part) > 1)


def _normalize_date_field(value: str) -> str:
    text = norm_text(value)
    if not text:
        return ""
    match = re.search(r"\b(\d{1,2}/\d{1,2}/\d{4}(?:\s+\d{1,2}:\d{2}\s+[AP]M)?)\b", text, flags=re.I)
    if match:
        return match.group(1)
    year_match = re.search(r"\b(20\d{2}|19\d{2})\b", text)
    if year_match:
        return year_match.group(1)
    return text


def _make_unstructured_pdf_record(entity_name: str, owner_meta: dict) -> dict:
    return {
        "_schedule": "PDF",
        "_record_type": "pdf_entity_unstructured",
        "entity_name": entity_name,
        "raw_value": entity_name,
        "owner_last_name": owner_meta["owner_last_name"],
        "owner_first_name": owner_meta["owner_first_name"],
        "owner_middle_name": owner_meta["owner_middle_name"],
        "owner_full_name": owner_meta["owner_full_name"],
        "filer_position_title": owner_meta.get("filer_position_title", ""),
        "filer_agency_name": owner_meta.get("filer_agency_name", ""),
        "filer_entity_name": owner_meta.get("filer_entity_name", ""),
        "jurisdiction": owner_meta.get("jurisdiction", ""),
        "_source_pdf_path": owner_meta.get("_source_pdf_path", ""),
    }


def _is_entity_candidate_line(value: str, owner_keys: set[str]) -> bool:
    text = _clean_pdf_candidate(value)
    key = norm_key(text)
    if not key or key in owner_keys:
        return False
    if not _is_valid_pdf_entity(text):
        return False
    if _is_contact_or_address_line(text):
        return False
    if _looks_like_description_line(text):
        return False
    if _looks_like_role_or_status_line(text):
        return False
    if _looks_like_person_name_line(text):
        return False
    return _entity_signal_score(text) >= 1


def _context_supports_entity_line(
    current: str,
    previous: str,
    following: str,
    owner_keys: set[str],
) -> bool:
    current = _clean_pdf_candidate(current)
    previous = _clean_pdf_candidate(previous)
    following = _clean_pdf_candidate(following)
    current_tokens = [token for token in re.findall(r"[A-Za-z0-9&'.-]+", current) if re.search(r"[A-Za-z]", token)]
    prev_key = norm_key(previous)
    next_key = norm_key(following)
    prev_desc = _looks_like_description_line(previous) or _looks_like_role_or_status_line(previous)
    next_desc = _looks_like_description_line(following) or _looks_like_role_or_status_line(following)
    prev_owner = prev_key in owner_keys
    next_owner = next_key in owner_keys
    prev_address = _is_contact_or_address_line(previous)
    next_address = _is_contact_or_address_line(following)
    has_business_signal = _has_business_or_institution_signal(current)

    if len(current_tokens) == 1 and not has_business_signal:
        return prev_desc or next_desc
    if not has_business_signal:
        return prev_desc or next_desc or (prev_owner and next_desc) or (next_owner and prev_desc)
    return prev_desc or next_desc or prev_owner or next_owner or prev_address or next_address or has_business_signal


def _extract_records_from_unstructured_pdf_text(text: str, owner_meta: dict) -> list[dict]:
    lines = _normalized_lines(text)
    if not lines:
        return []

    owner_keys = {
        norm_key(owner_meta.get("owner_full_name", "")),
        norm_key(owner_meta.get("owner_first_name", "")),
        norm_key(owner_meta.get("owner_last_name", "")),
    }
    owner_keys.discard("")

    start_index = 0
    for index, line in enumerate(lines):
        if re.search(r"\bsan:\s*fppc\b", line, flags=re.I) or re.search(r"\bfiled date\b", line, flags=re.I):
            start_index = index + 1

    records = []
    seen = set()
    for index in range(start_index, len(lines)):
        line = _clean_pdf_candidate(lines[index])
        key = norm_key(line)
        if not key or key in seen:
            continue
        if not _is_entity_candidate_line(line, owner_keys):
            continue

        previous = lines[index - 1] if index > 0 else ""
        following = lines[index + 1] if index + 1 < len(lines) else ""
        if not _context_supports_entity_line(line, previous, following, owner_keys):
            continue

        seen.add(key)
        records.append(_make_unstructured_pdf_record(line, owner_meta))
    return records


def extract_form700_metadata_from_text(
    text: str,
    filing_metadata: dict | None = None,
    source_pdf_path: str = "",
    pdf_title: str = "",
) -> dict:
    fallback = _owner_meta_from_filing_metadata(filing_metadata, source_pdf_path)
    lines = _normalized_lines(text)
    title_first, title_middle, title_last = _extract_name_from_pdf_title(pdf_title)

    filer_section = _extract_named_section_lines(
        lines,
        title_patterns=[r"filer"],
        required_patterns=[r"first name", r"last name|surname"],
        stop_patterns=[r"filing info", r"expanded statement list", r"filing data", r"gifts and travel"] + COMMON_SECTION_TERMINATORS,
    )
    filing_info_section = _extract_named_section_lines(
        lines,
        title_patterns=[r"filing info"],
        required_patterns=[r"filing type", r"filed date|filing year"],
        stop_patterns=[r"expanded statement list", r"filing data", r"gifts and travel"] + COMMON_SECTION_TERMINATORS,
    )

    first = _extract_line_labeled_value(filer_section, [r"first name"], PDF_FILER_LABELS) or title_first or fallback["owner_first_name"]
    middle = _extract_line_labeled_value(
        filer_section,
        [r"middle name", r"middle initial"],
        PDF_FILER_LABELS,
    ) or title_middle or fallback["owner_middle_name"]
    last = _extract_line_labeled_value(
        filer_section,
        [r"last name", r"surname"],
        PDF_FILER_LABELS,
    ) or title_last or fallback["owner_last_name"]
    agency = _extract_line_labeled_value(filer_section, [r"agency"], PDF_FILER_LABELS) or fallback.get("filer_agency_name", "")
    position = _extract_line_labeled_value(filer_section, [r"position"], PDF_FILER_LABELS) or fallback.get("filer_position_title", "")

    filing_type = _extract_line_labeled_value(filing_info_section, [r"filing type"], PDF_FILING_INFO_LABELS)
    filing_year = _extract_line_labeled_value(filing_info_section, [r"filing year"], PDF_FILING_INFO_LABELS)
    due_date = _extract_line_labeled_value(filing_info_section, [r"due date"], PDF_FILING_INFO_LABELS)
    filed_date = _extract_line_labeled_value(filing_info_section, [r"filed date"], PDF_FILING_INFO_LABELS)

    full_text_label_stops = PDF_FILER_LABELS + PDF_FILING_INFO_LABELS + PDF_SECTION_STOP_LINES + COMMON_SECTION_TERMINATORS
    filing_type = filing_type or _extract_text_labeled_value(text, [r"filing type"], full_text_label_stops)
    filing_year = filing_year or _extract_text_labeled_value(text, [r"filing year"], full_text_label_stops)
    due_date = due_date or _extract_text_labeled_value(text, [r"due date"], full_text_label_stops)
    filed_date = filed_date or _extract_text_labeled_value(text, [r"filed date"], full_text_label_stops)

    if not filing_year:
        year_match = re.search(r"\bfiling year\b\s*:?\s*(20\d{2}|19\d{2})\b", "\n".join(lines), flags=re.I)
        if year_match:
            filing_year = year_match.group(1)
    if not due_date:
        due_match = re.search(r"\bdue date\b\s*:?\s*(\d{1,2}/\d{1,2}/\d{4})\b", "\n".join(lines), flags=re.I)
        if due_match:
            due_date = due_match.group(1)
    if not filed_date:
        filed_match = re.search(r"\bfiled date\b\s*:?\s*(\d{1,2}/\d{1,2}/\d{4})\b", "\n".join(lines), flags=re.I)
        if filed_match:
            filed_date = filed_match.group(1)

    if not first or not last or not agency or not position:
        cover_first, cover_middle, cover_last, cover_position, cover_agency = _extract_unstructured_cover_owner_fields(
            lines,
            fallback,
        )
        first = first or cover_first
        middle = middle or cover_middle
        last = last or cover_last
        position = position or cover_position
        agency = agency or cover_agency

    if not agency or not position:
        unstructured_position, unstructured_agency = _extract_unstructured_header_fields(lines, fallback)
        position = position or unstructured_position
        agency = agency or unstructured_agency

    filing_year = _normalize_date_field(filing_year)
    due_date = _normalize_date_field(due_date)
    filed_date = _normalize_date_field(filed_date)

    full = " ".join(part for part in [first, middle, last] if part).strip() or fallback["owner_full_name"]
    return {
        "owner_last_name": last,
        "owner_first_name": first,
        "owner_middle_name": middle,
        "owner_full_name": full,
        "filer_position_title": position,
        "filer_agency_name": agency,
        "filer_entity_name": fallback.get("filer_entity_name", ""),
        "filing_type": filing_type,
        "filing_year": filing_year,
        "due_date": due_date,
        "filed_date": filed_date,
        "jurisdiction": fallback.get("jurisdiction", ""),
        "_source_pdf_path": source_pdf_path,
    }


def extract_form700_metadata_from_pdf(
    input_path: str | Path,
    filing_metadata: dict | None = None,
) -> dict:
    input_path = Path(input_path)
    text, title = extract_pdf_text_and_title(input_path)
    return extract_form700_metadata_from_text(
        text,
        filing_metadata=filing_metadata,
        source_pdf_path=str(input_path),
        pdf_title=title,
    )


def _split_schedule_sections(text: str) -> dict[str, list[str]]:
    markers = []
    for schedule, cfg in PDF_SCHEDULE_TARGETS.items():
        for pattern in cfg["section_patterns"]:
            for match in re.finditer(pattern, text, flags=re.I):
                markers.append((match.start(), schedule))

    markers.sort(key=lambda item: item[0])
    sections = {schedule: [] for schedule in PDF_SCHEDULE_TARGETS}
    if not markers:
        for schedule in sections:
            sections[schedule].append(text)
        return sections

    for index, (start, schedule) in enumerate(markers):
        end = markers[index + 1][0] if index + 1 < len(markers) else len(text)
        chunk = text[start:end]
        sections[schedule].append(chunk)
    return sections


def _extract_records_from_schedule_text(
    schedule: str,
    text: str,
    owner_meta: dict,
) -> list[dict]:
    cfg = PDF_SCHEDULE_TARGETS[schedule]
    all_terminators = list(cfg["terminators"]) + COMMON_SECTION_TERMINATORS
    terminator_pattern = "|".join(f"(?:{term})" for term in all_terminators)
    records = []
    seen = set()

    for label in cfg["capture_labels"]:
        pattern = re.compile(
            rf"{label}\s*[:\-]?\s*(.+?)(?=(?:{terminator_pattern})|$)",
            re.I | re.S,
        )
        for match in pattern.finditer(text):
            raw_value = _clean_pdf_candidate(match.group(1))
            if not _is_valid_pdf_entity(raw_value):
                continue
            key = (schedule, norm_key(raw_value))
            if key in seen:
                continue
            seen.add(key)
            records.append(
                {
                    "_schedule": schedule,
                    "_record_type": cfg["record_type"],
                    "entity_name": raw_value,
                    "raw_value": raw_value,
                    "owner_last_name": owner_meta["owner_last_name"],
                    "owner_first_name": owner_meta["owner_first_name"],
                    "owner_middle_name": owner_meta["owner_middle_name"],
                    "owner_full_name": owner_meta["owner_full_name"],
                    "filer_position_title": owner_meta.get("filer_position_title", ""),
                    "filer_agency_name": owner_meta.get("filer_agency_name", ""),
                    "filer_entity_name": owner_meta.get("filer_entity_name", ""),
                    "jurisdiction": owner_meta.get("jurisdiction", ""),
                    "_source_pdf_path": owner_meta.get("_source_pdf_path", ""),
                }
            )
    return records


def extract_form700_records_from_text(
    text: str,
    filing_metadata: dict | None = None,
    source_pdf_path: str = "",
    pdf_title: str = "",
) -> list[dict]:
    owner_meta = extract_form700_metadata_from_text(
        text,
        filing_metadata=filing_metadata,
        source_pdf_path=source_pdf_path,
        pdf_title=pdf_title,
    )
    normalized_text = text.replace("\r", "\n")
    sections = _split_schedule_sections(normalized_text)
    records = []
    for schedule, chunks in sections.items():
        for chunk in chunks:
            records.extend(_extract_records_from_schedule_text(schedule, chunk, owner_meta))
    if not records:
        records = _extract_records_from_unstructured_pdf_text(normalized_text, owner_meta)
    return records


def sanitize_form700_records(records: list[dict]) -> list[dict]:
    cleaned_records = []
    seen = set()
    for row in records:
        record_type = norm_key(row.get("_record_type", ""))
        owner_full_name = norm_text(row.get("owner_full_name", ""))
        if record_type == "pdfentityunstructured" and not owner_full_name:
            continue

        entity_name = _clean_pdf_candidate(row.get("entity_name", ""))
        if not _is_valid_pdf_entity(entity_name):
            continue
        dedupe_key = (
            norm_key(owner_full_name),
            record_type,
            norm_key(entity_name),
            norm_key(row.get("_source_pdf_path", "")),
        )
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        cleaned = dict(row)
        cleaned["entity_name"] = entity_name
        cleaned["filer_agency_name"] = _sanitize_context_field(row.get("filer_agency_name", ""))
        cleaned["filer_position_title"] = _sanitize_context_field(row.get("filer_position_title", ""))
        cleaned_records.append(cleaned)
    return cleaned_records


def parse_form700_pdf(input_path: str | Path, filing_metadata: dict | None = None) -> list[dict]:
    input_path = Path(input_path)
    print(f"[info] opening Form 700 PDF: {input_path}")
    text, title = extract_pdf_text_and_title(input_path)
    records = extract_form700_records_from_text(
        text,
        filing_metadata=filing_metadata,
        source_pdf_path=str(input_path),
        pdf_title=title,
    )
    records = sanitize_form700_records(records)
    print(f"[info] Form 700 PDF parsing complete: {len(records)} total rows")
    return records


def export_form700_database(
    database_path: str | Path,
    out_csv: str | Path,
    out_json: str | Path,
    jurisdiction: str = "",
) -> list[dict]:
    database = MinutesDatabase(database_path)
    database.initialize()
    rows = database.fetch_form700_entity_rows(jurisdiction=jurisdiction)
    rows = sanitize_form700_records(rows)
    write_outputs(rows, out_csv, out_json)
    return rows


def write_outputs(records, out_csv: str | Path, out_json: str | Path):
    out_csv = Path(out_csv)
    out_json = Path(out_json)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    out_json.parent.mkdir(parents=True, exist_ok=True)

    print(f"[info] writing Form 700 CSV output: {out_csv}")
    print(f"[info] writing Form 700 JSON output: {out_json}")

    fieldnames = []
    seen = set()
    for row in records:
        for k in row.keys():
            if k not in seen:
                seen.add(k)
                fieldnames.append(k)

    with out_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)

    out_json.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")
    print("[info] finished writing Form 700 outputs")
